from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
import os
import sqlite3
import json
from config import Config
from predict_engine import predict_churn, get_all_predictions

app = Flask(__name__)
app.config.from_object(Config)

# ── Create folders if they don't exist ──────────────────────────────────────
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['MODEL_FOLDER'], exist_ok=True)
os.makedirs(app.config['CHARTS_FOLDER'], exist_ok=True)

# ── Database setup ───────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS predictions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id TEXT,
            tenure INTEGER,
            monthly_charges REAL,
            total_charges REAL,
            contract_type TEXT,
            internet_service TEXT,
            payment_method TEXT,
            churn_prediction TEXT,
            churn_probability REAL,
            risk_badge TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS bulk_predictions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT,
            total_customers INTEGER,
            churn_count INTEGER,
            non_churn_count INTEGER,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id TEXT,
            feedback_text TEXT,
            rating INTEGER,
            sentiment TEXT,
            polarity REAL,
            subjectivity REAL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

# ── Routes ───────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

# ── Predict Route ─────────────────────────────────────────────────────────────
@app.route('/predict', methods=['GET', 'POST'])
def predict():
    if request.method == 'POST':
        customer_data = {
            'customer_id':      request.form.get('customer_id'),
            'gender':           request.form.get('gender', 'Male'),
            'senior_citizen':   request.form.get('senior_citizen', 0),
            'partner':          request.form.get('partner', 'No'),
            'dependents':       request.form.get('dependents', 'No'),
            'tenure':           request.form.get('tenure', 0),
            'phone_service':    request.form.get('phone_service', 'Yes'),
            'multiple_lines':   request.form.get('multiple_lines',
                                                  'No phone service'),
            'internet_service': request.form.get('internet_service', 'DSL'),
            'online_security':  request.form.get('online_security', 'No'),
            'online_backup':    request.form.get('online_backup', 'No'),
            'device_protection':request.form.get('device_protection', 'No'),
            'tech_support':     request.form.get('tech_support', 'No'),
            'streaming_tv':     request.form.get('streaming_tv', 'No'),
            'streaming_movies': request.form.get('streaming_movies', 'No'),
            'contract_type':    request.form.get('contract_type',
                                                  'Month-to-month'),
            'paperless_billing':request.form.get('paperless_billing', 'Yes'),
            'payment_method':   request.form.get('payment_method',
                                                  'Electronic check'),
            'monthly_charges':  request.form.get('monthly_charges', 0),
            'total_charges':    request.form.get('total_charges', 0),
            'num_products':     request.form.get('num_products', 1)
        }

        result = predict_churn(customer_data)

        if result['success']:
            return render_template('result.html', result=result)
        else:
            flash(f"Prediction error: {result['error']}", 'danger')
            return redirect(url_for('predict'))

    return render_template('predict.html')

# ── Dashboard Route ───────────────────────────────────────────────────────────
@app.route('/dashboard')
def dashboard():
    predictions = get_all_predictions()
    total       = len(predictions)
    high_risk   = sum(1 for p in predictions if p['risk_badge'] == 'High')
    low_risk    = sum(1 for p in predictions if p['risk_badge'] == 'Low')
    churn_rate  = round((high_risk / total * 100), 1) if total > 0 else 0
    return render_template('dashboard.html',
                           predictions=predictions,
                           total=total,
                           high_risk=high_risk,
                           low_risk=low_risk,
                           churn_rate=churn_rate)

# ── Bulk Upload Route ─────────────────────────────────────────────────────────
@app.route('/bulk_upload', methods=['GET', 'POST'])
def bulk_upload():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected!', 'danger')
            return redirect(url_for('bulk_upload'))

        file = request.files['file']

        if file.filename == '':
            flash('No file selected!', 'danger')
            return redirect(url_for('bulk_upload'))

        if not file.filename.endswith('.csv'):
            flash('Only CSV files are allowed!', 'danger')
            return redirect(url_for('bulk_upload'))

        try:
            import pandas as pd
            from predict_engine import predict_churn

            df           = pd.read_csv(file)
            results      = []
            churn_count  = 0
            total        = len(df)

            for _, row in df.iterrows():
                customer_data = {
                    'customer_id': row.get('customerID',
                                   row.get('customer_id', 'N/A')),
                    'gender':           row.get('gender', 'Male'),
                    'senior_citizen':   row.get('SeniorCitizen',
                                        row.get('senior_citizen', 0)),
                    'partner':          row.get('Partner',
                                        row.get('partner', 'No')),
                    'dependents':       row.get('Dependents',
                                        row.get('dependents', 'No')),
                    'tenure':           row.get('tenure', 0),
                    'phone_service':    row.get('PhoneService',
                                        row.get('phone_service', 'Yes')),
                    'multiple_lines':   row.get('MultipleLines',
                                        row.get('multiple_lines',
                                                'No phone service')),
                    'internet_service': row.get('InternetService',
                                        row.get('internet_service', 'DSL')),
                    'online_security':  row.get('OnlineSecurity',
                                        row.get('online_security', 'No')),
                    'online_backup':    row.get('OnlineBackup',
                                        row.get('online_backup', 'No')),
                    'device_protection':row.get('DeviceProtection',
                                        row.get('device_protection', 'No')),
                    'tech_support':     row.get('TechSupport',
                                        row.get('tech_support', 'No')),
                    'streaming_tv':     row.get('StreamingTV',
                                        row.get('streaming_tv', 'No')),
                    'streaming_movies': row.get('StreamingMovies',
                                        row.get('streaming_movies', 'No')),
                    'contract_type':    row.get('Contract',
                                        row.get('contract_type',
                                                'Month-to-month')),
                    'paperless_billing':row.get('PaperlessBilling',
                                        row.get('paperless_billing', 'Yes')),
                    'payment_method':   row.get('PaymentMethod',
                                        row.get('payment_method',
                                                'Electronic check')),
                    'monthly_charges':  row.get('MonthlyCharges',
                                        row.get('monthly_charges', 0)),
                    'total_charges':    row.get('TotalCharges',
                                        row.get('total_charges', 0))
                }

                result = predict_churn(customer_data)
                if result['success']:
                    results.append(result)
                    if result['risk_badge'] == 'High':
                        churn_count += 1

            conn   = sqlite3.connect(app.config['DATABASE_PATH'])
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO bulk_predictions
                (filename, total_customers, churn_count, non_churn_count)
                VALUES (?, ?, ?, ?)
            ''', (file.filename, total, churn_count, total - churn_count))
            conn.commit()
            conn.close()

            return render_template('bulk_result.html',
                                   results=results,
                                   total=total,
                                   churn_count=churn_count,
                                   filename=file.filename)

        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'danger')
            return redirect(url_for('bulk_upload'))

    return render_template('bulk_upload.html')

# ── Reports Route ─────────────────────────────────────────────────────────────
@app.route('/reports')
def reports():
    predictions = get_all_predictions()
    return render_template('reports.html', predictions=predictions)

# ── Export PDF Route ──────────────────────────────────────────────────────────
@app.route('/export/pdf')
def export_pdf():
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table,
                                    TableStyle, Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet
    from flask import Response
    import io

    predictions = get_all_predictions()
    buffer      = io.BytesIO()
    doc         = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles      = getSampleStyleSheet()
    elements    = []

    elements.append(Paragraph(
        "ChurnPredict AI — Prediction Report", styles['Title']))
    elements.append(Spacer(1, 20))

    data = [['#', 'Customer ID', 'Tenure',
             'Monthly $', 'Contract', 'Risk', 'Probability', 'Time']]

    for i, p in enumerate(predictions[:50], 1):
        data.append([
            str(i),
            str(p['customer_id']),
            str(p['tenure']),
            f"${p['monthly_charges']}",
            str(p['contract_type']),
            str(p['risk_badge']),
            f"{p['churn_probability']}%",
            str(p['timestamp'])[:16]
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#667eea')),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 11),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1),
         [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID',        (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE',    (0,1), (-1,-1), 9),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition':
                 'attachment; filename=churn_report.pdf'}
    )

# ── Export Excel Route ────────────────────────────────────────────────────────
@app.route('/export/excel')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import Response
    import io

    predictions = get_all_predictions()
    wb          = Workbook()
    ws          = wb.active
    ws.title    = 'Churn Predictions'

    header_fill = PatternFill(start_color='667eea',
                               end_color='667eea',
                               fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['#', 'Customer ID', 'Tenure', 'Monthly Charges',
               'Total Charges', 'Contract', 'Internet Service',
               'Payment Method', 'Risk Badge',
               'Churn Probability', 'Timestamp']
    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')

    for i, p in enumerate(predictions, 2):
        ws.cell(row=i, column=1,  value=i-1)
        ws.cell(row=i, column=2,  value=p['customer_id'])
        ws.cell(row=i, column=3,  value=p['tenure'])
        ws.cell(row=i, column=4,  value=p['monthly_charges'])
        ws.cell(row=i, column=5,  value=p['total_charges'])
        ws.cell(row=i, column=6,  value=p['contract_type'])
        ws.cell(row=i, column=7,  value=p['internet_service'])
        ws.cell(row=i, column=8,  value=p['payment_method'])
        ws.cell(row=i, column=9,  value=p['risk_badge'])
        ws.cell(row=i, column=10, value=p['churn_probability'])
        ws.cell(row=i, column=11, value=str(p['timestamp']))

        risk_cell = ws.cell(row=i, column=9)
        if p['risk_badge'] == 'High':
            risk_cell.fill = PatternFill(start_color='FFE0E0',
                                          end_color='FFE0E0',
                                          fill_type='solid')
        elif p['risk_badge'] == 'Medium':
            risk_cell.fill = PatternFill(start_color='FFF3CD',
                                          end_color='FFF3CD',
                                          fill_type='solid')
        else:
            risk_cell.fill = PatternFill(start_color='D4EDDA',
                                          end_color='D4EDDA',
                                          fill_type='solid')

    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[
            col[0].column_letter].width = max_length + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument'
                 '.spreadsheetml.sheet',
        headers={'Content-Disposition':
                 'attachment; filename=churn_report.xlsx'}
    )

# ── Export CSV Route ──────────────────────────────────────────────────────────
@app.route('/export/csv')
def export_csv():
    import csv
    import io
    from flask import Response

    predictions = get_all_predictions()
    output      = io.StringIO()
    writer      = csv.writer(output)

    writer.writerow(['#', 'Customer ID', 'Tenure', 'Monthly Charges',
                     'Total Charges', 'Contract', 'Internet Service',
                     'Payment Method', 'Risk Badge',
                     'Churn Probability', 'Timestamp'])

    for i, p in enumerate(predictions, 1):
        writer.writerow([
            i, p['customer_id'], p['tenure'],
            p['monthly_charges'], p['total_charges'],
            p['contract_type'], p['internet_service'],
            p['payment_method'], p['risk_badge'],
            p['churn_probability'], p['timestamp']
        ])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition':
                 'attachment; filename=churn_report.csv'}
    )

# ── Feedback Route ────────────────────────────────────────────────────────────
@app.route('/feedback', methods=['GET', 'POST'])
def feedback():
    if request.method == 'POST':
        try:
            from textblob import TextBlob
            customer_id   = request.form.get('customer_id', 'Anonymous')
            feedback_text = request.form.get('feedback_text', '')
            rating        = int(request.form.get('rating', 3))

            blob         = TextBlob(feedback_text)
            polarity     = blob.sentiment.polarity
            subjectivity = blob.sentiment.subjectivity

            if polarity > 0.2:
                sentiment       = 'Positive'
                sentiment_color = 'success'
                emoji           = '😊'
            elif polarity < -0.2:
                sentiment       = 'Negative'
                sentiment_color = 'danger'
                emoji           = '😞'
            else:
                sentiment       = 'Neutral'
                sentiment_color = 'warning'
                emoji           = '😐'

            conn = sqlite3.connect(app.config['DATABASE_PATH'])
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO feedback
                (customer_id, feedback_text, rating,
                 sentiment, polarity, subjectivity)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (customer_id, feedback_text, rating,
                  sentiment, round(polarity, 3),
                  round(subjectivity, 3)))
            conn.commit()
            conn.close()

            result = {
                'customer_id':     customer_id,
                'feedback_text':   feedback_text,
                'rating':          rating,
                'sentiment':       sentiment,
                'sentiment_color': sentiment_color,
                'emoji':           emoji,
                'polarity':        round(polarity, 3),
                'subjectivity':    round(subjectivity, 3)
            }
            return render_template('feedback.html',
                                   result=result,
                                   submitted=True,
                                   feedbacks=[])

        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
            return redirect(url_for('feedback'))

    try:
        conn = sqlite3.connect(app.config['DATABASE_PATH'])
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM feedback ORDER BY timestamp DESC"
        )
        feedbacks = cursor.fetchall()
        conn.close()
    except:
        feedbacks = []

    return render_template('feedback.html',
                           submitted=False,
                           feedbacks=feedbacks,
                           result=None)

# ── About Route ───────────────────────────────────────────────────────────────
@app.route('/about')
def about():
    return render_template('about.html')

# ── Run the app ──────────────────────────────────────────────────────────────
if __name__ == '__main__':
    init_db()
    print("✅ Database initialized successfully!")
    print("🚀 Starting Customer Churn Prediction App...")
    print("🌐 Open your browser and go to: http://127.0.0.1:5000")
    app.run(debug=True)